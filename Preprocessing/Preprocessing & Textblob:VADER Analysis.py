#!/usr/bin/env python
# -*- coding:utf-8 -*-
#!usr/bin/env python
#coding:utf-8


import csv
from textblob import TextBlob
import re
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

from nltk.stem.porter import PorterStemmer
from nltk.tokenize import word_tokenize
import openpyxl
def ClearExcel(filename):   #如果表格有旧的数据 先清空
    try:
      wb = openpyxl.load_workbook(filename)   # 数据写入excel
    except:
      return
    wb = openpyxl.load_workbook(filename)  # 数据写入excel

    sheet = wb.active
    print(sheet.title)
    wb.create_sheet()
    wb.remove(sheet)

    wb.save(filename)
    #sheet = wb.active
    #print(sheet.title)


def WriteToExcel(tasklist,filename,column_names = None,deleteId=False):

    try:
      wb = openpyxl.load_workbook(filename)   # 数据写入excel
    except:
      wb = openpyxl.Workbook()
      wb.save(filename)
    sheet = wb.active
    rows = sheet.max_row
    i = 1
    for every_task in tasklist:
        coln = 1
        #print(type(every_task))
        if(isinstance(every_task,str)):
            sheet.cell(row=1, column=1).value = 'text'
            sheet.cell(rows + i, 1).value = str(every_task)
        elif(isinstance(every_task,list)):
            for e in every_task:
                #print(e)
                sheet.cell(rows + i, coln).value = e
                sheet.cell(row=1, column=coln).value = 'a'
                coln+=1
        else:
            for key in every_task.__dict__:
                if deleteId == True:
                    if key =='ID':
                        continue
                value = every_task.__dict__[key]
                sheet.cell(rows + i, coln).value = str(value)
                sheet.cell(row = 1, column =coln).value = key
                coln += 1

        i = i+1
    if column_names != None:
        coln = 1
        for name in column_names:
            sheet.cell(row=1, column=coln).value = name
            coln += 1
    wb.save(filename)





def AnalyzeCsvFile(filename,column = 2):

    i = 0
    taskList = []
    with open(filename, 'r',encoding='utf8') as f:
        read = csv.reader(f)
        for line in read:
            #print(line)
            try:
                i += 1
                ID = 0
                sentence = line[column-1]
                if sentence =='tweet' or i==1:
                    continue
                # 替换掉非法字符，防止openpyxl 报openpyxl.utils.exceptions.IllegalCharacterError错误
                ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                sentence = ILLEGAL_CHARACTERS_RE.sub(r'', sentence)
                sentence = clean_tweet(sentence)
                polarity,subjectivity,vader_neg,vader_neu,vader_pos,vader_compound = analize_sentiment(sentence)
                print(sentence, polarity,subjectivity,vader_neg,vader_neu,vader_pos,vader_compound)
                taskList.append([sentence,polarity,subjectivity,vader_neg,vader_neu,vader_pos,vader_compound])

                if line[1]=='' and line[10]=='':
                    break
            except:
                pass
    print(i)

    # 到这里就是EXCEL分析完成了
    ClearExcel('分析结果1.xlsx')
    header = ['tweet','polarity','subjectivity','vader_neg','vader_neu','vader_pos','vader_compound']
    #WriteToExcel(taskList, '分析结果1.xlsx',column_names =header, deleteId=False)
    spl = filename.split('/')
    newpath = spl[len(spl) - 1]
    newpath = newpath.split('.')[0]
    newpath = mkdir('result/' + newpath)
    ClearExcel(newpath + '/' + '分析结果1.xlsx')
    WriteToExcel(taskList, newpath + '/' + '分析结果1.xlsx',column_names =header, deleteId=False)
    # sqlSaveOne(filename,newpath)
    print(newpath)
    return taskList, newpath + '/'
def stopwords(sentence):
    lst = []
    with open ('stopwords_EN.txt','r',encoding='utf8') as f:
        pass
        lines = f.readlines()
        for l in lines:
            lst.append(l.strip('\n').strip())
        f.close()
    for l in lst:
        sentence = sentence.replace(l,'')
    return sentence
def loadStopWords():
    lst = []
    with open ('stopwords_EN.txt','r',encoding='utf8') as f:
        pass
        lines = f.readlines()
        for l in lines:
            lst.append(l.strip('\n').strip())
        f.close()
    return lst
def isStopWords(word,vols):

    for l in vols:
        if word.lower().strip() == l.lower().strip():
            return True
    return False
def ProcessCsvFile(filename,column = 2):

    i = 0
    taskList = []
    with open(filename, 'r',encoding='utf8') as f:
        read = csv.reader(f)
        for line in read:
            #print(line)
            try:
                i += 1

                sentence = line[column-1]
                if sentence =='tweet' or i==1:
                    continue
                # 替换掉非法字符，防止openpyxl 报openpyxl.utils.exceptions.IllegalCharacterError错误
                ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                sentence = ILLEGAL_CHARACTERS_RE.sub(r'', sentence)
                sentence = clean_tweet(sentence)
                #sentence = stopwords(sentence)
                input_text = sentence.split(' ')
                #print(input_text)
                # 去除停用词
                vols = loadStopWords()
                new_lst = []
                for t in input_text:
                    if not isStopWords(t,vols):
                        new_lst.append(t)
                input_text = new_lst
                # 词干分析
                porter = PorterStemmer()
                output = ' '
                for word in input_text:
                    r = porter.stem(word)
                    output = output +r+' '
                #print('after stem',output)
                #print('tokenize:')
                # tokenization
                tok = word_tokenize(output)
                print(i,' tokenize:',tok)
                taskList.append(tok)
                if line[1]=='' and line[10]=='':
                    break
            except:
                pass
    print(i)


    # 到这里是EXCEL分析完成
    ClearExcel('processed.xlsx')
    header = ['tweet','polarity','subjectivity','vader_neg','vader_neu','vader_pos','vader_compound']
    #WriteToExcel(taskList, '分析结果1.xlsx',column_names =header, deleteId=False)
    spl = filename.split('/')
    newpath = spl[len(spl) - 1]
    newpath = newpath.split('.')[0]
    newpath = mkdir('result/' + newpath)
    ClearExcel(newpath + '/' + 'processed.xlsx')
    WriteToExcel(taskList, newpath + '/' + 'processed.xlsx', deleteId=False)
    # sqlSaveOne(filename,newpath)
    print(newpath)
    return taskList, newpath + '/'

def clean_tweet(tweet):

    return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)", " ", tweet).split())

def analize_sentiment(tweet):

    analysis = TextBlob(clean_tweet(tweet))
    #sentiment_analyzer_scores(clean_tweet(tweet))
    # Analyze using vader
    analyser = SentimentIntensityAnalyzer()
    vader_score = analyser.polarity_scores(clean_tweet(tweet))
    print(vader_score)
    return analysis.sentiment.polarity,analysis.sentiment.subjectivity,vader_score['neg'],vader_score['neu'],vader_score['pos'],vader_score['compound']




def mkdir(path):
    # 引入模块
    import os

    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    #path = path.rstrip("\\")
    opath = os.getcwd()
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)

    # 判断结果
    if not isExists:
        # 如果不存在则创建目录

        os.makedirs(path)
        print(path + ' 创建成功')


    else:
        # 如果目录存在则不创建，并提示目录已存在
        #print(path + ' 目录已存在')
        pass


    return str(opath)+'/'+path



if __name__ == '__main__':
    import nltk
    nltk.download('punkt')
    #s = 'i got id'
    #print(word_tokenize(s))
    ProcessCsvFile('vaccination_tweets.csv',column = 11)
    #AnalyzeCsvFile('vaccination_tweets.csv',column = 11)
